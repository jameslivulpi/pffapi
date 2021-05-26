#!/usr/bin/env python3

from datetime import date, timedelta
import time
from copy import deepcopy
import os
import requests
from openpyxl import Workbook

class API:
    def __init__(self):
        self.api_key = os.environ.get("PFF_API_KEY")
        self.base_url = "https://api.profootballfocus.com/v1/"
        self.jwt_token = ""
        self.headers = ""
        self.row = 1
        self.sheet, self.wb = self._PrepareExcel()


    def _PrepareExcel(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "NFL"

        #create titles for each column
        sheet.cell(row=1, column=1).value='Player ID'
        sheet.cell(row=1, column=2).value='Player Name'
        sheet.cell(row=1, column=3).value='player Age'
        sheet.cell(row=1, column=4).value='Position'
        sheet.cell(row=1, column=5).value='Draft Year'
        sheet.cell(row=1, column=6).value='Draft Round'
        sheet.cell(row=1, column=7).value='Draft Selection'


        sheet.cell(row=1, column=8).value='2020 Seasonal Offense Grade'
        sheet.cell(row=1, column=9).value='2020 Seasonal Defense Grade'
        sheet.cell(row=1, column=10).value='2020 Total Snap Count'

        sheet.cell(row=1, column=11).value='2019 Seasonal Offense Grade'
        sheet.cell(row=1, column=12).value='2019 Seasonal Defense Grade'
        sheet.cell(row=1, column=13).value='2019 Total Snap Count'

        sheet.cell(row=1, column=14).value='2020 Seasonal Offense Grade (NCAA)'
        sheet.cell(row=1, column=15).value='2020 Seasonal Defense Grade (NCAA)'
        sheet.cell(row=1, column=16).value='2019 Seasonal Offense Grade (NCAA)'
        sheet.cell(row=1, column=17).value='2019 Seasonal Defense Grade (NCAA)'

        return sheet, wb
    @staticmethod
    def _CalcAge(birthdate):
        if birthdate is None:
            return None
        age = (date.today() - date.fromisoformat(birthdate)) // timedelta(days=365.2425)
        return age

    def GetPosition(self, playerid):
        self.headers = {'User-Agent': "hello","Authorization": "Bearer " + self.jwt_token}

        res = requests.get(self.base_url+"nfl"+"/players/latest/", headers=self.headers, params = {"id": playerid})
        if res.status_code == 200:
            print(res.status_code)
            try:
                for item in res.json()['rosters']:
                    self.sheet.cell(row=self.row+1, column=4).value=item['position'] if "position" in item else "NA"
                    for pick in item['drafts'] or item['draft']:
                        self.sheet.cell(row=self.row+1, column=5).value=pick['season']
                        if pick['season'] == 2021:
                            self.sheet.cell(row=self.row+1, column=6).value=pick['round']
                            self.sheet.cell(row=self.row+1, column=7).value=pick['selection']
                            self.GetGrade(playerid, "ncaa", 2020)
                            self.GetGrade(playerid, "ncaa", 2019)
            except:
                print(f"no info for {playerid}")


        if res.status_code == 429:
            print(res.status_code)
            time.sleep(5)
            self.GetPosition(playerid)


    def GetGrade(self, playerid, league,  year):
        self.headers = {'User-Agent': "hello","Authorization": "Bearer " + self.jwt_token}
        res = requests.get(self.base_url+f"grades/{league}/{year}/season_grade/", headers=self.headers)
        if res.status_code == 200:
            tmp_list = []
            try:
                for grade in res.json()['season_grade']:
                    if playerid == grade['player_id']:
                        try:
                            if self.sheet.cell(row=self.row+1, column=4).value == "NA" or self.sheet.cell(row=self.row+1, column=4).value is None:
                                self.sheet.cell(row=self.row+1, column=4).value = grade['position']
                        except:
                            pass
                    else:
                        continue
            except:
                print("no season grade info for: ", playerid)
                return

            if len(tmp_list) == 0:
                return

            _max_dict = {}
            _max = 0
            tmp_list.sort(key=lambda x:x['week'], reverse=False)
            for x in tmp_list:
                if x['week'] >= _max:
                    _max_dict = deepcopy(x)

            try:
                if year == 2020 and league == "nfl":
                    self.sheet.cell(row=self.row+1, column=8).value=_max_dict['offense']
                if year == 2019 and league == "nfl":
                    self.sheet.cell(row=self.row+1, column=11).value=_max_dict['offense']
                 #NCAA
                if year == 2020 and league == "ncaa":
                    self.sheet.cell(row=self.row+1, column=14).value=_max_dict['offense']
                if year == 2019 and league == "ncaa":
                    self.sheet.cell(row=self.row+1, column=16).value=_max_dict['offense']
            except:
                if year == 2020 and league == "nfl":
                    self.sheet.cell(row=self.row+1, column=8).value="N/A"
                if year == 2019 and league == "nfl":
                    self.sheet.cell(row=self.row+1, column=11).value="N/A"

                if year == 2020 and league == "ncaa":
                    self.sheet.cell(row=self.row+1, column=14).value="N/A"
                if year == 2019 and league == "ncaa":
                    self.sheet.cell(row=self.row+1, column=16).value="N/A"



            try:
                if year == 2020 and league == "nfl":
#                   print(_maddx_dict['offense_rank'])
                    self.sheet.cell(row=self.row+1, column=9).value=_max_dict['defense']
                if year == 2019 and league == "nfl":
                    self.sheet.cell(row=self.row+1, column=12).value=_max_dict['defense']
                if year == 2020 and league == "ncaa":
                    self.sheet.cell(row=self.row+1, column=15).value=_max_dict['defense']
                if year == 2019 and league == "ncaa":
                    self.sheet.cell(row=self.row+1, column=17).value=_max_dict['defense']
            except:
                if year == 2020 and league == "nfl":
                    self.sheet.cell(row=self.row+1, column=9).value="N/A"
                if year == 2019 and league == "nfl":
                    self.sheet.cell(row=self.row+1, column=12).value="N/A"
                if year == 2020 and league == "ncaa":
                    self.sheet.cell(row=self.row+1, column=15).value="N/A"
                if year == 2019 and league == "ncaa":
                    self.sheet.cell(row=self.row+1, column=17).value="N/A"

            if year == 2020 and league == "nfl" or (league == "ncaa" and  self.sheet.cell(row=self.row+1, column=12).value is None):
                self.sheet.cell(row=self.row+1, column=10).value=_max_dict['total_snaps']
            if year == 2019 and league == "nfl" or (league == "ncaa" and  self.sheet.cell(row=self.row+1, column=17).value is None):
                self.sheet.cell(row=self.row+1, column=13).value=_max_dict['total_snaps']

        elif res.status_code == 401:
            self.login()
            #retry current:
            self.GetGrade(playerid, league,  year)



        else:
            print(f"GetGrade returns status code {res.status_code}")
            time.sleep(5)
            self.GetGrade(playerid, league,  year)




    def login(self):
        login_header = {"x-api-key": self.api_key}
        res = requests.post("https://api.profootballfocus.com/auth/login", headers = login_header)
        if res.status_code == 200:
            print("Login success")
            self.jwt_token = res.json()['jwt']


    def run(self):
        self.headers = {'User-Agent': "bot 0.1","Authorization": "Bearer " + self.jwt_token}
        res = requests.get(self.base_url+"nfl"+"/players/", headers=self.headers, params = {"page_size": 10000000})
        if res.status_code == 200:
            for item in res.json()['players']:
                if not "retired" in item:
                    self.sheet.cell(row=self.row+1, column=1).value=item['id']
                    self.sheet.cell(row=self.row+1, column=2).value=item['first_name'] + " " + item['last_name']
                    age = self._CalcAge(item['dob'])
                    self.sheet.cell(row=self.row+1, column=3).value=age
                    self.GetPosition(item['id'])
                    self.GetGrade(int(item['id']),"nfl", 2020)
                    self.GetGrade(int(item['id']), "nfl",  2019)
                    self.row += 1
                    self.wb.save("nfl-bobby-test.xlsx")




if __name__ == "__main__":
    x = API()
    x.login()
#    #x.GetPosition(57905)
#    x.GetGrade(143797, "ncaa", 2019)
    x.run()
